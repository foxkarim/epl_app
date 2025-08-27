# app.py — EPL Excel → Streamlit (trim numbers, CF approx, custom views, PLYDEX team builder with club cap)

import os
import re
import math
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ---------------------- App Config ----------------------
st.set_page_config(page_title="EPL Workbook Browser", layout="wide")

# Hide these sheets from the UI (case-insensitive names)
HIDDEN_SHEETS = {"index"}  # hide the Index sheet
DEFAULT_WORKBOOK = "EPL macro.xlsm"

st.title("EPL Macro Workbook → Web App")
st.caption("Custom sheet views, **trimmed numeric display** (no extra zeros), Excel static styles, approximate conditional formatting, and a PLYDEX-based team builder with club cap (max 3 per club).")

# ---------------------- Utilities ----------------------
def normalize_name(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def find_col(df: pd.DataFrame, targets):
    """Find first matching column (case-insensitive) from a list of names/synonyms."""
    if not isinstance(targets, (list, tuple)): targets = [targets]
    lower = {c.lower(): c for c in df.columns}
    norm = {normalize_name(c): c for c in df.columns}
    for t in targets:
        t_norm = normalize_name(t)
        if t.lower() in lower:
            return lower[t.lower()]
        if t_norm in norm:
            return norm[t_norm]
    # substring fallback
    for t in targets:
        t_low = t.lower()
        for c in df.columns:
            if t_low in c.lower():
                return c
    return None

def match_columns(df: pd.DataFrame, wanted_with_synonyms: list):
    """wanted_with_synonyms = list of lists, each inner list: [preferred, *synonyms]"""
    out = []
    for syns in wanted_with_synonyms:
        col = find_col(df, syns)
        if col is not None:
            out.append(col)
    return out

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")

# ---- Global "trim number" formatting ----
def fmt_number(x):
    """Return a minimal string for numbers (no trailing zeros, no scientific notation)."""
    try:
        if pd.isna(x):
            return ""
        xf = float(x)
        if math.isfinite(xf):
            if xf.is_integer():
                return str(int(xf))
            # avoid scientific notation, trim trailing zeros
            return np.format_float_positional(xf, trim='-')
    except Exception:
        pass
    return x

def style_trim_numbers(styler):
    """Apply fmt_number as a global formatter to a Styler."""
    return styler.format(fmt_number)

def df_with_trim(df: pd.DataFrame):
    """Return a Styler that formats all numbers trimmed."""
    return df.style.format(fmt_number)

# ---------------------- Load data ----------------------
@st.cache_data(show_spinner=False)
def load_excel_all_sheets(path_or_file) -> dict:
    dfs = pd.read_excel(path_or_file, sheet_name=None, engine="openpyxl")
    cleaned = {}
    for name, df in dfs.items():
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        cleaned[name] = df
    return cleaned

# ---------------------- Excel static styles helpers ----------------------
def _argb_to_hex(argb: str) -> str:
    if not argb:
        return ""
    argb = str(argb)
    if len(argb) == 8:
        return "#" + argb[-6:]
    if len(argb) == 6:
        return "#" + argb
    return ""

def read_sheet_styles(xlsm_path: str, sheet_name: str):
    wb = load_workbook(xlsm_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    vals, css = [], []
    max_row, max_col = ws.max_row, ws.max_column

    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_vals, row_css = [], []
        for cell in r:
            row_vals.append(cell.value)

            bg = ""
            fill = getattr(cell, "fill", None)
            if fill:
                rgb = getattr(fill.fgColor, "rgb", None) or getattr(fill.start_color, "rgb", None)
                if rgb and rgb != "00000000":
                    bg = _argb_to_hex(rgb)

            f = getattr(cell, "font", None)
            css_parts = []
            if bg:
                css_parts.append(f"background-color:{bg}")
            if f:
                if f.color and getattr(f.color, "rgb", None):
                    css_parts.append(f"color:{_argb_to_hex(f.color.rgb)}")
                if f.bold:
                    css_parts.append("font-weight:bold")
                if f.italic:
                    css_parts.append("font-style:italic")
                if f.underline:
                    css_parts.append("text-decoration: underline")

            row_css.append("; ".join(css_parts))
        vals.append(row_vals)
        css.append(row_css)

    values_df = pd.DataFrame(vals)

    # header heuristic
    header_row = None
    if len(vals) >= 1:
        first = vals[0]
        non_empty = any(v is not None and str(v).strip() != "" for v in first)
        unique = len(set(first)) == len(first)
        if non_empty and unique:
            header_row = 0

    if header_row == 0:
        values_df.columns = [str(c) if c is not None else "" for c in values_df.iloc[0]]
        values_df = values_df.iloc[1:].reset_index(drop=True)
        css_df = pd.DataFrame(css).iloc[1:].reset_index(drop=True)
        css_df.columns = values_df.columns
    else:
        css_df = pd.DataFrame(css)

    return values_df, css_df

def style_dataframe(values_df, css_df):
    css_df = css_df.reindex_like(values_df).fillna("")
    styler = values_df.style
    def _apply(_df):
        return css_df
    styler = styler.apply(_apply, axis=None)
    styler = style_trim_numbers(styler)  # trim numbers
    styler = styler.set_properties(**{"white-space": "nowrap"})
    return styler

# ---------------------- CF approximation helpers ----------------------
def a1_to_col_indices(a1_range):
    m = re.match(r"\$?([A-Z]+)\$?(\d+)?:\$?([A-Z]+)\$?(\d+)?", a1_range)
    if m:
        c1, _, c2, _ = m.groups()
        i1 = column_index_from_string(c1) - 1
        i2 = column_index_from_string(c2) - 1
        return list(range(min(i1, i2), max(i1, i2)+1))
    m2 = re.match(r"\$?([A-Z]+):\$?([A-Z]+)", a1_range)
    if m2:
        c1, c2 = m2.groups()
        i1 = column_index_from_string(c1) - 1
        i2 = column_index_from_string(c2) - 1
        return list(range(min(i1, i2), max(i1, i2)+1))
    m3 = re.match(r"\$?([A-Z]+)\$?\d+", a1_range)
    if m3:
        return [column_index_from_string(m3.group(1)) - 1]
    return []

def parse_cf(ws):
    out = []
    cf = ws.conditional_formatting
    for rng in getattr(cf, "_cf_rules", {}):
        for rule in cf._cf_rules[rng]:
            kind = getattr(rule, "type", None) or getattr(rule, "rule_type", None)
            if kind == "colorScale":
                cs = getattr(rule, "colorScale", None)
                if not cs: continue
                thresholds = []
                for cfvo, color in zip(cs.cfvo, cs.color):
                    thresholds.append({
                        "type": cfvo.type,
                        "val": cfvo.val,
                        "color": _argb_to_hex(getattr(color, "rgb", "")),
                    })
                out.append({"kind":"colorScale","range":rng,"thresholds":thresholds})
            elif kind == "cellIs":
                op = getattr(rule, "operator", None)
                dxf = getattr(rule, "dxf", None)
                bg = None
                if dxf and getattr(dxf, "fill", None):
                    col = getattr(dxf.fill.fgColor, "rgb", None) or getattr(dxf.fill.start_color, "rgb", None)
                    if col: bg = _argb_to_hex(col)
                f_list = getattr(rule, "formula", None) or []
                out.append({"kind":"cellIs","range":rng,"operator":op,"formula":f_list,"bg":bg})
    return out

def apply_cf_colors(df: pd.DataFrame, ws):
    rules = parse_cf(ws)
    if not rules: return None
    headers = [cell.value for cell in ws[1]]
    styler = df.style
    def affected_cols(a1):
        ws_cols = a1_to_col_indices(a1)
        cols = []
        if len(headers) == len(df.columns):
            for c in ws_cols:
                if 0 <= c < len(df.columns):
                    cols.append(df.columns[c])
        else:
            for c in ws_cols:
                if 0 <= c < len(headers):
                    hv = headers[c]
                    if hv is not None and str(hv).strip() in df.columns:
                        cols.append(str(hv).strip())
        seen, ordered = set(), []
        for x in cols:
            if x in df.columns and x not in seen:
                seen.add(x); ordered.append(x)
        return ordered
    for r in rules:
        if r["kind"]=="colorScale":
            cols = []
            for part in str(r["range"]).split():
                cols += affected_cols(part)
            for col in cols:
                if pd.api.types.is_numeric_dtype(df[col]):
                    styler = styler.background_gradient(axis=None, subset=[col])
    def mask_series(series, operator, formula_vals):
        s = pd.to_numeric(series, errors="coerce")
        try:
            if operator in ("lessThan","lt"): return s < float(formula_vals[0])
            if operator in ("lessThanOrEqual","le"): return s <= float(formula_vals[0])
            if operator in ("greaterThan","gt"): return s > float(formula_vals[0])
            if operator in ("greaterThanOrEqual","ge"): return s >= float(formula_vals[0])
            if operator in ("equal","eq"): return s == float(formula_vals[0])
            if operator == "between": return s.between(float(formula_vals[0]), float(formula_vals[1]), inclusive="both")
        except: return pd.Series(False, index=series.index)
        return pd.Series(False, index=series.index)
    for r in rules:
        if r["kind"]=="cellIs" and r.get("bg"):
            cols = []
            for part in str(r["range"]).split():
                cols += affected_cols(part)
            for col in cols:
                if col in df.columns:
                    m = mask_series(df[col], r.get("operator"), r.get("formula", []))
                    css_col = pd.Series([""]*len(df), index=df.index, dtype="object")
                    css_col[m.fillna(False)] = f"background-color:{r['bg']}"
                    styler = styler.apply(lambda s: css_col, subset=[col])
    styler = style_trim_numbers(styler)  # trim numbers in CF view
    return styler

# ---------------------- Custom sheet renderers ----------------------
# Colors
YELLOW = "#FFF2CC"
LIGHT_GREEN_BG = "#C6EFCE"
LIGHT_GRAY = "#D9D9D9"
GRAY = "#BFBFBF"
RED = "#FF0000"
DARK_RED = "#8B0000"
DARK_GREEN = "#006400"
GOLD = "#FFD700"
DARK_GOLD = "#A67C00"
WHITE = "#FFFFFF"
BLUE = "#0000FF"
LIGHT_BLUE = "#CFE2F3"
BLACK = "#000000"
GRAY_FONT = "#7F7F7F"

def render_schedule(df: pd.DataFrame):
    wanted = [
        ["Game Week", "GW", "Gameweek"],
        ["Schedule"],
        ["Home Team", "Home"],
        ["Home Goals", "HG"],
        ["Time"],
        ["Away Goals", "AG"],
        ["Away Team", "Away"],
    ]
    cols = match_columns(df, wanted)
    if not cols:
        st.info("Couldn't find expected Schedule columns; showing sheet as-is.")
        st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
    else:
        st.dataframe(df[cols].style.format(fmt_number), use_container_width=True, hide_index=True)

def render_faceoff(df: pd.DataFrame):
    # numbers: 1–19 => yellow; 20–38 => light green
    def color_cell(val):
        try:
            x = float(val)
        except:
            return ""
        if 1 <= x <= 19:
            return f"background-color:{YELLOW}"
        if 20 <= x <= 38:
            return f"background-color:{LIGHT_GREEN_BG}"
        return ""
    styler = df.style.applymap(color_cell)
    styler = style_trim_numbers(styler)
    st.dataframe(styler, use_container_width=True)

def render_fdrating(df: pd.DataFrame):
    # 1..5 mapping
    def style_val(v):
        try:
            n = int(float(v))
        except:
            return ""
        if n == 1:
            return f"font-weight:bold; color:{GOLD}; background-color:{DARK_GREEN}"
        if n == 2:
            return f"font-weight:bold; color:{BLACK}; background-color:{LIGHT_GREEN_BG}"
        if n == 3:
            return f"font-weight:bold; color:{BLUE}; background-color:{GRAY}"
        if n == 4:
            return f"font-weight:bold; color:{GRAY_FONT}; background-color:{RED}"
        if n == 5:
            return f"font-weight:bold; color:{WHITE}; background-color:{DARK_RED}"
        return ""
    styler = df.style.applymap(style_val)
    styler = style_trim_numbers(styler)
    st.dataframe(styler, use_container_width=True)

def render_table_sheet(df: pd.DataFrame):
    # Hide helper column(s) in TableRank table
    drop_cols = [c for c in df.columns if "helper" in c.lower() or "tablerank helper" in c.lower() or c.strip().startswith("_")]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)

def render_plydex(df: pd.DataFrame):
    # First & last column styled per Position; ONLY the Position column is uppercased
    pos_col = find_col(df, ["Position", "Pos"])
    if pos_col is None:
        st.info("Position column not found; showing sheet as-is.")
        st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
        return
    if df.shape[1] < 2:
        st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
        return
    first_col = df.columns[0]
    last_col = df.columns[-1]

    df_display = df.copy()
    # Only uppercase the Position column
    df_display[pos_col] = df_display[pos_col].astype(str).str.upper()

    def css_for_pos(val_upper: str):
        val = (val_upper or "").strip().upper()
        if val == "GKP":
            return f"font-weight:bold; color:{BLACK}; background-color:{LIGHT_GREEN_BG}"
        if val == "DEF":
            return f"font-weight:bold; color:{DARK_GOLD}; background-color:{LIGHT_GRAY}"
        if val == "MID":
            return f"font-weight:bold; color:{WHITE}; background-color:{RED}"
        if val == "FWD":
            return f"font-weight:bold; color:{BLACK}; background-color:{LIGHT_BLUE}"
        return ""

    # Build per-cell CSS: apply to first_col, last_col, and pos_col itself
    def build_css_df(_df):
        css_rows = []
        for i, row in _df.iterrows():
            css = ["" for _ in _df.columns]
            style_str = css_for_pos(row[pos_col])
            for idx, c in enumerate(_df.columns):
                if c in (first_col, last_col, pos_col):
                    css[idx] = style_str
            css_rows.append(css)
        return pd.DataFrame(css_rows, index=_df.index, columns=_df.columns)

    styler = df_display.style.apply(build_css_df, axis=None)
    styler = style_trim_numbers(styler)
    st.dataframe(styler, use_container_width=True)

# ---------------------- PLYDEX-based Team Builder (with club cap) ----------------------
ROSTER_SLOTS = (
    ["GKP1","GKP2"] +
    [f"DEF{i}" for i in range(1,6)] +
    [f"MID{i}" for i in range(1,6)] +
    [f"FWD{i}" for i in range(1,4)]
)
SLOT_TO_POS = {s: s[:3] for s in ROSTER_SLOTS}  # GKP/DEF/MID/FWD

def tm_find_cols(df: pd.DataFrame):
    return {
        "player": find_col(df, ["Player", "Name", "Player Name"]),
        "pos": find_col(df, ["Position", "Pos"]),
        "club": find_col(df, ["Club", "Team"]),
        "value": find_col(df, ["Value", "Price", "Market Value", "Transfer Value", "Cost"]),
        "nation": find_col(df, ["Nationality", "Nation", "Country"]),
    }

def init_team_slots():
    if "team_slots" not in st.session_state:
        st.session_state.team_slots = {s: None for s in ROSTER_SLOTS}
    if "selected_slot" not in st.session_state:
        st.session_state.selected_slot = ROSTER_SLOTS[0]

def team_df_from_slots():
    rows = []
    for slot, rec in st.session_state.team_slots.items():
        if rec is None: continue
        rows.append({
            "Slot": slot,
            "Player": rec.get("Player",""),
            "Position": rec.get("Position",""),
            "Club": rec.get("Club",""),
            "Value": rec.get("Value",""),
        })
    return pd.DataFrame(rows, columns=["Slot","Player","Position","Club","Value"])

def club_counts():
    """Return dict of club(lowercased)->count across all filled slots."""
    counts = {}
    for rec in st.session_state.team_slots.values():
        if not rec: 
            continue
        club = str(rec.get("Club", "")).strip()
        if not club:
            continue
        key = club.lower()
        counts[key] = counts.get(key, 0) + 1
    return counts

def render_team_builder_from_plydex(dfs: dict):
    st.markdown("### Build & Save Your 15‑Player Team (from PLYDEX → Matrix) — Max 3 per club") 

    # Find PLYDEX sheet
    plydex_key = None
    for k in dfs.keys():
        if "plydex" in k.lower():
            plydex_key = k
            break
    if plydex_key is None:
        st.error("Couldn't find the PLYDEX sheet. Please ensure it exists.")
        st.dataframe(df_with_trim(dfs[list(dfs.keys())[0]]), use_container_width=True, hide_index=True)
        return

    src = dfs[plydex_key]
    cols = tm_find_cols(src)
    if cols["player"] is None or cols["pos"] is None:
        st.error("PLYDEX sheet must have at least Player and Position columns.")
        st.dataframe(df_with_trim(src), use_container_width=True, hide_index=True)
        return

    init_team_slots()

    # ---- Slot grid ----
    st.caption("Click a slot below to fill it. The search table will auto-filter by that slot's position. (Club cap: 3)")
    grid_cols = st.columns(5)
    for idx, slot in enumerate(ROSTER_SLOTS):
        col = grid_cols[idx % 5]
        with col:
            filled = st.session_state.team_slots[slot]
            label = f"**{slot}**" if st.session_state.selected_slot == slot else slot
            if filled:
                label += f"\n{filled['Player']}"
            if st.button(label, key=f"slot_{slot}"):
                st.session_state.selected_slot = slot

    sel_slot = st.session_state.selected_slot
    sel_pos = SLOT_TO_POS[sel_slot]
    st.write(f"Selected slot: **{sel_slot}**  |  Position filter: **{sel_pos}**") 

    # ---- Search within PLYDEX by position ----
    df_pos = src[src[cols["pos"]].astype(str).str.upper() == sel_pos].copy()
    search = st.text_input("Search by player/club/nation (filtered to selected position)").strip().lower()
    if search:
        txt_cols = [c for c in [cols["player"], cols["club"], cols["nation"]] if c is not None]
        if txt_cols:
            mask = pd.Series(False, index=df_pos.index)
            for c in txt_cols:
                mask = mask | df_pos[c].astype(str).str.lower().str.contains(re.escape(search), na=False)
            df_pos = df_pos[mask]

    show_cols = [c for c in [cols["player"], cols["pos"], cols["club"], cols["value"]] if c]
    st.dataframe(df_pos[show_cols].head(300).style.format(fmt_number), use_container_width=True, hide_index=True)

    options = df_pos[cols["player"]].astype(str).unique().tolist()
    pick = st.selectbox("Select player for this slot", options)
    c1, c2 = st.columns([1,1])
    with c1:
        if st.button("➕ Add to selected slot"):
            # prevent duplicates
            already = {rec["Player"] for rec in st.session_state.team_slots.values() if rec}
            if pick in already:
                st.warning(f"{pick} is already in your team.")
            else:
                row = df_pos[df_pos[cols["player"]].astype(str) == pick].iloc[0]
                club_name = str(row.get(cols["club"], "")) if cols["club"] else ""
                key = club_name.strip().lower()
                counts = club_counts()
                if club_name and counts.get(key, 0) >= 3:
                    st.warning(f"You already have 3 players from {club_name}.")
                else:
                    st.session_state.team_slots[sel_slot] = {
                        "Player": str(row.get(cols["player"], "")),
                        "Position": str(row.get(cols["pos"], "")).upper(),
                        "Club": club_name,
                        "Value": row.get(cols["value"], ""),
                    }
                    st.success(f"Added {pick} to {sel_slot}.")
    with c2:
        if st.button("🗑️ Clear selected slot"):
            st.session_state.team_slots[sel_slot] = None
            st.info(f"Cleared {sel_slot}.")

    # ---- Current team & save/load ----
    st.markdown("#### Current Team")
    tdf = team_df_from_slots()
    if tdf.empty:
        st.info("No players yet. Click a slot and add a player.")
    else:
        st.dataframe(tdf.style.format(fmt_number), use_container_width=True, hide_index=True)
        st.download_button("Download team (CSV)", data=to_csv_bytes(tdf), file_name="my_team.csv")

    with st.expander("Load a saved team (CSV with columns Slot,Player,Position,Club,Value)"):
        up = st.file_uploader("Upload team CSV", type=["csv"], key="team_csv_loader")
        if up is not None:
            try:
                loaded = pd.read_csv(up)
                # reset slots then fill with club cap enforcement
                st.session_state.team_slots = {s: None for s in ROSTER_SLOTS}
                skipped = []
                counts = {}
                for _, r in loaded.iterrows():
                    slot = str(r.get("Slot",""))
                    club_name = str(r.get("Club",""))
                    key = club_name.strip().lower()
                    if club_name and counts.get(key, 0) >= 3:
                        skipped.append((slot, r.get("Player",""), club_name))
                        continue
                    if slot in st.session_state.team_slots:
                        st.session_state.team_slots[slot] = {
                            "Player": str(r.get("Player","")),
                            "Position": str(r.get("Position","")),
                            "Club": club_name,
                            "Value": r.get("Value",""),
                        }
                        if club_name:
                            counts[key] = counts.get(key, 0) + 1
                if skipped:
                    st.warning("Some entries were skipped to enforce max 3 per club: " + ", ".join([f"{p} ({c}) @ {s}" for s,p,c in skipped]))
                else:
                    st.success("Team loaded into slots.")
            except Exception as e:
                st.error(f"Could not read team CSV: {e}")

# ---------------------- Sidebar ----------------------
with st.sidebar:
    st.header("Workbook")
    uploaded = st.file_uploader("Upload Excel (.xlsx / .xlsm)", type=["xlsx", "xlsm"])

    if uploaded is not None:
        dfs = load_excel_all_sheets(uploaded)
        source_label = "Uploaded file"
        local_path_for_file = None
    else:
        if os.path.exists(DEFAULT_WORKBOOK):
            dfs = load_excel_all_sheets(DEFAULT_WORKBOOK)
            source_label = f"Local file: {DEFAULT_WORKBOOK}"
            local_path_for_file = DEFAULT_WORKBOOK
        else:
            st.info("Upload an Excel file to get started.")
            st.stop()

    st.success(f"Loaded: {source_label}")

    # Build sheet list excluding hidden
    sheet_names = [s for s in dfs.keys() if s and s.strip().lower() not in HIDDEN_SHEETS]
    sheet = st.selectbox("Sheet", sorted(sheet_names))

    mode = st.radio(
        "View mode",
        ["Filterable", "Static Excel formatting", "Conditional formatting (approx)"],
        index=0,
        help="Static shows saved fills/fonts. CF approximates color scales + simple numeric 'Cell Is' rules."
    )

# ---------------------- Main ----------------------
df = dfs[sheet].copy()
sheet_key = sheet.strip().lower()

st.subheader(f"Sheet: {sheet}")
st.write(f"Rows: **{len(df):,}** | Columns: **{len(df.columns)}**")

# Custom routing for particular sheets (in Filterable mode)
if mode == "Filterable":
    if "schedule" in sheet_key:
        render_schedule(df); st.stop()
    if "face" in sheet_key and "off" in sheet_key:  # Face-Off
        render_faceoff(df); st.stop()
    if "fd" in sheet_key and "rating" in sheet_key:
        render_fdrating(df); st.stop()
    if sheet_key == "table":
        render_table_sheet(df); st.stop()
    if "plydex" in sheet_key:
        render_plydex(df); st.stop()
    if "transfer" in sheet_key:
        # Use PLYDEX-based team builder
        render_team_builder_from_plydex(dfs); st.stop()

# ---------- Static Excel formatting ----------
if mode == "Static Excel formatting":
    if uploaded is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp:
            tmp.write(uploaded.getbuffer())
            path_for_styles = tmp.name
    else:
        path_for_styles = local_path_for_file
    try:
        vals_df, css_df = read_sheet_styles(path_for_styles, sheet)
        styler = style_dataframe(vals_df, css_df)
        st.dataframe(styler, use_container_width=True)
        st.caption("Static Excel formatting (solid fills + font styles)." )
    except Exception as e:
        st.warning(f"Could not render static styles: {e}")
        st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
    st.stop()

# ---------- Conditional Formatting (approx) ----------
if mode == "Conditional formatting (approx)":
    if uploaded is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp:
            tmp.write(uploaded.getbuffer())
            path_for_cf = tmp.name
    else:
        path_for_cf = local_path_for_file
    try:
        wb = load_workbook(path_for_cf, data_only=True)
        ws = wb[sheet]
        styler = apply_cf_colors(df, ws)
        if styler is None:
            st.info("No supported conditional formatting rules found (supports color scales + simple numeric comparisons)." )
            st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
        else:
            st.dataframe(styler, use_container_width=True)
            st.caption("Approximate rendering of Excel CF colors." )
    except Exception as e:
        st.warning(f"Could not apply conditional formatting: {e}")
        st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
    st.stop()

# ---------- Default filterable view (trim numbers) ----------
st.dataframe(df_with_trim(df), use_container_width=True, hide_index=True)
